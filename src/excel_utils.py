#!/usr/bin/env python3
"""Excel processing utilities"""

import pandas as pd
import openpyxl


def _load_workbook_and_get_merged_ranges(file_path: str, sheet_name: str):
    """Load workbook and get merged cell ranges"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    merged_ranges = list(ws.merged_cells.ranges)
    return wb, ws, merged_ranges


def get_merged_cells(file_path: str, sheet_name: str):
    """Get merged cells information"""
    _, ws, merged_ranges = _load_workbook_and_get_merged_ranges(file_path, sheet_name)
    
    cells = []
    for merged_range in merged_ranges:
        if merged_range.min_row == 6:  # Only process merged cells in row 6
            cell_value = ws.cell(merged_range.min_row, merged_range.min_col).value
            if cell_value and str(cell_value).strip():
                cells.append({
                    'value': str(cell_value).strip(),
                    'start_col': merged_range.min_col,
                    'end_col': merged_range.max_col + 1
                })
    
    return cells


def read_merged_headers(file_path: str, sheet_name: str) -> pd.DataFrame:
    """Read Excel file with merged headers"""
    _, ws, merged_ranges = _load_workbook_and_get_merged_ranges(file_path, sheet_name)
    
    column_names = []
    for col in range(1, ws.max_column + 1):
        parts = []
        for row in [7]:
            merged_cell = next((r for r in merged_ranges 
                              if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col), None)
            
            if merged_cell:
                cell_value = ws.cell(merged_cell.min_row, merged_cell.min_col).value
            else:
                cell_value = ws.cell(row, col).value
            
            if cell_value and str(cell_value).strip():
                part = str(cell_value).strip()
                if part not in parts:
                    parts.append(part)
        
        column_names.append(" - ".join(parts) if parts else f"Column_{col}")
    
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, skiprows=7)
    df.columns = column_names[:len(df.columns)]
    return df